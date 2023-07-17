#!/usr/bin/python

import requests
import time
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook

def scrape_website():
    # Send a GET request to the website
    response = requests.get('https://www.loyolaalumni.org/')
    content = response.content

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')

    # Extract image links from the <img> tags and prepend base URL
    img_elements = soup.find_all('img')
    img_links = [urljoin('https://www.loyolaalumni.org/', img.get('src')) for img in img_elements]

    # Extract hyperlink tags and their URLs from the <a> tags and prepend base URL
    a_elements = soup.find_all('a')
    hrefs = [urljoin('https://www.loyolaalumni.org/', a.get('href')) if a.get('href') else '' for a in a_elements]

    return img_links, hrefs

# Set the desired refresh interval in seconds
refresh_interval = 5

# Create a new Excel workbook
workbook = Workbook()

# Get the active sheet
sheet = workbook.active

# Set column headers
sheet['A1'] = 'Image Links'
sheet['B1'] = 'Hyperlink URLs'

# Run the loop indefinitely
while True:
    # Call the scraping function
    img_links, hrefs = scrape_website()

    # Populate the sheet with the extracted data
    for i in range(len(img_links)):
        sheet.cell(row=i+2, column=1, value=img_links[i])
        sheet.cell(row=i+2, column=2, value=hrefs[i])

    # Save the data to an Excel file
    workbook.save('data.xlsx')

    # Wait for the specified refresh interval
    time.sleep(refresh_interval)

